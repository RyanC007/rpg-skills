#!/usr/bin/env python3
"""
Technical SEO Audit & Fixes Spreadsheet Generator
Creates a professional Excel template for technical SEO audits
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_technical_seo_spreadsheet(filename="Technical_SEO_Audit_Template.xlsx"):
    """Generate a professional Technical SEO Audit spreadsheet"""
    
    wb = Workbook()
    
    # Define color scheme
    colors = {
        'header': 'FF1F4E78',  # Dark blue
        'critical': 'FFFF0000',  # Red
        'high': 'FFFF9900',  # Orange
        'medium': 'FFFFFF00',  # Yellow
        'low': 'FF92D050',  # Green
        'white': 'FFFFFFFF'
    }
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== SHEET 1: Executive Summary =====
    ws1 = wb.active
    ws1.title = "Executive Summary"
    
    # Headers
    summary_headers = [
        "Metric", "Value", "Notes"
    ]
    
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample data rows
    summary_data = [
        ["Total Issues Found", "=SUM('Critical Issues (P1)'!A:A,'High Priority (P2)'!A:A,'Medium Priority (P3)'!A:A,'Low Priority (P4)'!A:A)-4", "Total across all priority levels"],
        ["Critical Issues (P1)", "=COUNTA('Critical Issues (P1)'!A:A)-1", "Issues requiring immediate attention"],
        ["High Priority Issues (P2)", "=COUNTA('High Priority (P2)'!A:A)-1", "Issues to fix within 2-4 weeks"],
        ["Medium Priority Issues (P3)", "=COUNTA('Medium Priority (P3)'!A:A)-1", "Issues to fix within 1-3 months"],
        ["Low Priority Issues (P4)", "=COUNTA('Low Priority (P4)'!A:A)-1", "Issues to fix when possible"],
        ["", "", ""],
        ["Estimated Total Hours to Fix", "[Calculate based on individual issues]", "Sum of all estimated fix times"],
        ["Quick Wins (High Impact, Low Effort)", "[List top 5 quick wins]", "Prioritize these for immediate action"],
        ["", "", ""],
        ["Site Health Score", "[Calculate: 100 - (P1*5 + P2*3 + P3*1)]", "Higher is better (max 100)"],
        ["Expected Impact of Fixes", "[Describe expected improvements]", "Traffic increase, ranking improvements, etc."]
    ]
    
    for row_num, row_data in enumerate(summary_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 40
    ws1.column_dimensions['C'].width = 50
    
    ws1.freeze_panes = 'A2'
    
    # ===== SHEET 2-5: Issue Tracking Sheets =====
    priority_levels = [
        ("Critical Issues (P1)", colors['critical']),
        ("High Priority (P2)", colors['high']),
        ("Medium Priority (P3)", colors['medium']),
        ("Low Priority (P4)", colors['low'])
    ]
    
    issue_headers = [
        "Issue ID", "Issue Type", "Page/URL Affected", "Issue Description",
        "SEO Impact", "Business Impact", "Fix Instructions",
        "Estimated Time", "Required Skills", "Dependencies",
        "Status", "Notes"
    ]
    
    issue_col_widths = [12, 20, 40, 40, 15, 15, 50, 15, 20, 20, 15, 30]
    
    for sheet_name, priority_color in priority_levels:
        ws = wb.create_sheet(title=sheet_name)
        
        # Headers
        for col_num, header in enumerate(issue_headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Sample data row
        sample_data = [
            f"{sheet_name[:2]}-001",
            "[Issue Type]",
            "[URL]",
            "[Detailed description of the issue]",
            "[High/Medium/Low]",
            "[High/Medium/Low]",
            "[Step-by-step fix instructions]",
            "[X hours]",
            "[Required skills]",
            "[Dependencies]",
            "Not Started",
            "[Additional notes]"
        ]
        
        for col_num, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        # Column widths
        for col_num, width in enumerate(issue_col_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        ws.freeze_panes = 'A2'
    
    # ===== SHEET 6: Implementation Roadmap =====
    ws6 = wb.create_sheet(title="Implementation Roadmap")
    
    roadmap_headers = [
        "Week/Month", "Phase", "Issues to Address", "Estimated Hours",
        "Required Skills", "Expected Outcome", "Success Metrics", "Status", "Notes"
    ]
    
    for col_num, header in enumerate(roadmap_headers, 1):
        cell = ws6.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample roadmap data
    roadmap_data = [
        ["Week 1-2", "Critical Fixes", "P1-001 to P1-015", "10-15 hours", "HTML, CMS", "Site is crawlable and indexable", "Reduced errors in GSC", "Not Started", ""],
        ["Week 3-4", "High Priority", "P2-001 to P2-030", "15-20 hours", "HTML, Schema", "Key pages optimized", "Improved rankings", "Not Started", ""],
        ["Month 2-3", "Medium Priority", "P3-001 to P3-050", "20-30 hours", "Content, HTML", "All pages optimized", "Increased traffic", "Not Started", ""],
        ["Month 4-6", "Low Priority", "P4-001 to P4-025", "10-20 hours", "Various", "Site fully polished", "Sustained growth", "Not Started", ""]
    ]
    
    for row_num, row_data in enumerate(roadmap_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws6.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Column widths
    roadmap_col_widths = [15, 20, 30, 15, 20, 30, 25, 15, 30]
    for col_num, width in enumerate(roadmap_col_widths, 1):
        ws6.column_dimensions[get_column_letter(col_num)].width = width
    
    ws6.freeze_panes = 'A2'
    
    # ===== SHEET 7: Fix Instructions Library =====
    ws7 = wb.create_sheet(title="Fix Instructions Library")
    
    library_headers = [
        "Issue Type", "Description", "Why It Matters", "How to Fix", "Verification Steps", "Est. Time", "Skills Needed"
    ]
    
    for col_num, header in enumerate(library_headers, 1):
        cell = ws7.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample fix instructions
    fix_instructions = [
        [
            "Missing Title Tag",
            "Page is missing a <title> tag in the <head> section",
            "Title tags are critical for SEO and appear in search results",
            "1. Access page in CMS\\n2. Add <title> tag with 50-60 characters\\n3. Include primary keyword\\n4. Save and publish",
            "View page source and confirm <title> tag is present",
            "5-10 min",
            "Basic HTML or CMS"
        ],
        [
            "404 Error",
            "Page returns a 404 Not Found error",
            "Broken links harm user experience and waste crawl budget",
            "1. Identify all links pointing to this URL\\n2. Update links to correct URL or remove\\n3. Set up 301 redirect if appropriate\\n4. Verify in browser",
            "Check that URL no longer returns 404 and redirects work",
            "15-30 min",
            "HTML, Server config"
        ],
        [
            "Redirect Chain",
            "URL redirects multiple times before reaching destination",
            "Slows page load, wastes crawl budget, dilutes link equity",
            "1. Identify full redirect chain\\n2. Update initial redirect to point directly to final URL\\n3. Update internal links to bypass redirects",
            "Use redirect checker to confirm single redirect",
            "15-30 min",
            "Server config"
        ],
        [
            "Missing Meta Description",
            "Page is missing a meta description tag",
            "Meta descriptions appear in search results and affect CTR",
            "1. Access page in CMS\\n2. Add <meta name='description'> tag with 150-160 characters\\n3. Make it compelling and include keywords\\n4. Save and publish",
            "View page source and confirm meta description is present",
            "5-10 min",
            "Basic HTML or CMS"
        ],
        [
            "Missing H1 Tag",
            "Page is missing an H1 heading tag",
            "H1 tags help search engines understand page topic",
            "1. Access page in CMS or HTML\\n2. Add <h1> tag with primary keyword\\n3. Ensure only one H1 per page\\n4. Save and publish",
            "View page source and confirm H1 tag is present",
            "5 min",
            "Basic HTML or CMS"
        ]
    ]
    
    for row_num, row_data in enumerate(fix_instructions, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws7.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Column widths
    library_col_widths = [20, 30, 35, 50, 35, 12, 20]
    for col_num, width in enumerate(library_col_widths, 1):
        ws7.column_dimensions[get_column_letter(col_num)].width = width
    
    ws7.freeze_panes = 'A2'
    
    # Save the workbook
    wb.save(filename)
    print(f"✅ Technical SEO Audit spreadsheet created: {filename}")
    return filename

if __name__ == "__main__":
    create_technical_seo_spreadsheet()
