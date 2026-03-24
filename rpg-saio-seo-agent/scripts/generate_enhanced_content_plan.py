#!/usr/bin/env python3
"""
Enhanced Content Plan Spreadsheet Generator
Creates a professional Excel template for 6-month content planning with competitor analysis
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_enhanced_content_plan(filename="Enhanced_Content_Plan_Template.xlsx"):
    """Generate a professional Enhanced Content Plan spreadsheet"""
    
    wb = Workbook()
    
    # Define color scheme
    colors = {
        'header': 'FF1F4E78',  # Dark blue
        'accent': 'FF4472C4',  # Light blue
        'white': 'FFFFFFFF'
    }
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== SHEET 1: Master Content Plan =====
    ws1 = wb.active
    ws1.title = "Master Content Plan"
    
    master_headers = [
        "Page ID", "Page Type", "Page Name", "URL Slug", "Target Keywords (Primary)",
        "Target Keywords (Secondary)", "Word Count", "Priority", "Phase",
        "Content Type", "Schema Markup", "Internal Links TO", "Internal Links FROM",
        "Status", "Notes"
    ]
    
    for col_num, header in enumerate(master_headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample data
    sample_pages = [
        ["P001", "Homepage", "Home", "/", "primary keyword", "secondary keywords", "800-1000", "P1", "Phase 1", "Landing", "Organization, LocalBusiness", "All pages", "Nav menu", "Not Started", ""],
        ["P002", "Service", "Service Name", "/services/service-name", "service keyword", "related keywords", "1500-2000", "P1", "Phase 1", "Service Page", "Service", "Homepage, Related Services", "Nav, Footer", "Not Started", ""],
        ["P003", "About", "About Us", "/about", "company name", "about keywords", "800-1200", "P2", "Phase 1", "About Page", "Organization", "Homepage", "Nav, Footer", "Not Started", ""]
    ]
    
    for row_num, row_data in enumerate(sample_pages, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Column widths
    master_col_widths = [10, 15, 25, 25, 30, 30, 12, 10, 12, 15, 25, 30, 30, 15, 30]
    for col_num, width in enumerate(master_col_widths, 1):
        ws1.column_dimensions[get_column_letter(col_num)].width = width
    
    ws1.freeze_panes = 'A2'
    
    # ===== SHEET 2: Cornerstone Content =====
    ws2 = wb.create_sheet(title="Cornerstone Content")
    
    cornerstone_headers = [
        "Cornerstone Topic", "Main Page", "Supporting Content (Cluster)",
        "Target Keywords", "Word Count", "Internal Links", "Priority", "Notes"
    ]
    
    for col_num, header in enumerate(cornerstone_headers, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample cornerstone data
    cornerstone_data = [
        ["Main Service Topic", "/services/main-service", "Blog Post 1, Blog Post 2, FAQ Page", "primary keywords", "2000-3000", "All service pages link here", "P1", "Core pillar"],
        ["Secondary Topic", "/resources/secondary-topic", "Blog Post 3, Blog Post 4, Guide", "secondary keywords", "2000-2500", "Related service pages", "P2", "Supporting pillar"]
    ]
    
    for row_num, row_data in enumerate(cornerstone_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Column widths
    cornerstone_col_widths = [25, 30, 50, 30, 15, 35, 10, 30]
    for col_num, width in enumerate(cornerstone_col_widths, 1):
        ws2.column_dimensions[get_column_letter(col_num)].width = width
    
    ws2.freeze_panes = 'A2'
    
    # ===== SHEET 3: Linking Strategy =====
    ws3 = wb.create_sheet(title="Linking Strategy")
    
    linking_headers = [
        "Page", "Page Type", "Priority Links TO This Page",
        "Priority Links FROM This Page", "Anchor Text Examples", "Linking Notes"
    ]
    
    for col_num, header in enumerate(linking_headers, 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample linking data
    linking_data = [
        ["Homepage", "Landing", "All nav pages, Footer pages", "All cornerstone pages, Top service pages", "Learn more about [service], Explore [topic]", "Hub page - links to all important pages"],
        ["Service Page 1", "Service", "Homepage, Related services", "Blog posts about this service, FAQ", "Our [service] process, How we [service]", "Cornerstone page - receives many internal links"]
    ]
    
    for row_num, row_data in enumerate(linking_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Column widths
    linking_col_widths = [25, 15, 40, 40, 40, 40]
    for col_num, width in enumerate(linking_col_widths, 1):
        ws3.column_dimensions[get_column_letter(col_num)].width = width
    
    ws3.freeze_panes = 'A2'
    
    # ===== SHEET 4: Implementation Timeline =====
    ws4 = wb.create_sheet(title="Implementation Timeline")
    
    timeline_headers = [
        "Phase", "Timeline", "Page/Content", "Priority", "Dependencies",
        "Estimated Hours", "Assigned To", "Status", "Notes"
    ]
    
    for col_num, header in enumerate(timeline_headers, 1):
        cell = ws4.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample timeline data
    timeline_data = [
        ["Phase 1: Foundation", "Weeks 1-4", "Homepage, About, Contact", "P1", "None", "20-30 hours", "[Team Member]", "Not Started", "Core pages first"],
        ["Phase 1: Foundation", "Weeks 1-4", "Service Page 1, Service Page 2", "P1", "Homepage complete", "30-40 hours", "[Team Member]", "Not Started", "Key service pages"],
        ["Phase 2: Expansion", "Weeks 5-12", "Blog Posts 1-8", "P2", "Service pages complete", "40-60 hours", "[Team Member]", "Not Started", "Content creation phase"],
        ["Phase 3: Authority", "Weeks 13-24", "Blog Posts 9-24, Resources", "P2/P3", "Foundation complete", "60-80 hours", "[Team Member]", "Not Started", "Scale content production"]
    ]
    
    for row_num, row_data in enumerate(timeline_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Column widths
    timeline_col_widths = [20, 15, 30, 10, 25, 15, 20, 15, 30]
    for col_num, width in enumerate(timeline_col_widths, 1):
        ws4.column_dimensions[get_column_letter(col_num)].width = width
    
    ws4.freeze_panes = 'A2'
    
    # ===== SHEET 5: Blog Content Calendar (Enhanced) =====
    ws5 = wb.create_sheet(title="Blog Content Calendar")
    
    blog_headers = [
        "Month", "Week", "Blog Post Title", "Target Keywords", "Content Type",
        "Word Count", "Links To", "Priority", "Competitor Gap", "Competitive Advantage",
        "Est. Traffic Potential", "Content Depth", "Status"
    ]
    
    for col_num, header in enumerate(blog_headers, 1):
        cell = ws5.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample blog calendar data
    blog_data = [
        ["Month 1", "Week 1", "How to [Topic]: Complete Guide", "how to [topic]", "How-To Guide", "2000-2500", "Service Page 1", "P1", "Yes", "More comprehensive than competitors", "High", "Comprehensive", "Not Started"],
        ["Month 1", "Week 2", "Top 10 [Topic] Tips for [Audience]", "topic tips", "Listicle", "1500-2000", "Service Page 1", "P2", "No", "Better examples and visuals", "Medium", "Intermediate", "Not Started"],
        ["Month 1", "Week 3", "[Topic] Cost Guide: What to Expect", "topic cost", "Cost Guide", "1800-2200", "Service Page 2", "P1", "Yes", "Only guide with local pricing", "High", "Comprehensive", "Not Started"],
        ["Month 1", "Week 4", "Case Study: How We Helped [Client]", "case study", "Case Study", "1200-1500", "Service Page 1", "P2", "Yes", "Real data and results", "Medium", "Advanced", "Not Started"]
    ]
    
    for row_num, row_data in enumerate(blog_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws5.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Column widths
    blog_col_widths = [10, 10, 40, 25, 15, 12, 20, 10, 15, 35, 18, 15, 15]
    for col_num, width in enumerate(blog_col_widths, 1):
        ws5.column_dimensions[get_column_letter(col_num)].width = width
    
    ws5.freeze_panes = 'A2'
    
    # ===== SHEET 6: Competitor Content Matrix (NEW) =====
    ws6 = wb.create_sheet(title="Competitor Content Matrix")
    
    competitor_headers = [
        "Competitor Name", "Last 10 Post Titles", "Publication Dates",
        "Avg Frequency (posts/month)", "Content Types Used", "Topics Covered",
        "Avg Word Count", "Engagement Indicators", "Strengths", "Weaknesses", "Gaps"
    ]
    
    for col_num, header in enumerate(competitor_headers, 1):
        cell = ws6.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample competitor data
    competitor_data = [
        ["Competitor A", "[List of 10 titles]", "[Dates]", "8", "How-To (60%), Listicles (30%)", "Topic 1, Topic 2, Topic 3", "800", "High comments", "Consistent posting", "Superficial content", "No cost guides"],
        ["Competitor B", "[List of 10 titles]", "[Dates]", "4", "Listicles (70%), News (30%)", "Topic 1, Topic 4", "600", "Low engagement", "Good visuals", "Outdated info", "No local content"]
    ]
    
    for row_num, row_data in enumerate(competitor_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws6.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Column widths
    competitor_col_widths = [20, 50, 30, 22, 30, 30, 15, 25, 25, 25, 30]
    for col_num, width in enumerate(competitor_col_widths, 1):
        ws6.column_dimensions[get_column_letter(col_num)].width = width
    
    ws6.freeze_panes = 'A2'
    
    # ===== SHEET 7: Content Gap Opportunities (NEW) =====
    ws7 = wb.create_sheet(title="Content Gap Opportunities")
    
    gap_headers = [
        "Gap ID", "Gap Type", "Description", "Related Keywords",
        "Est. Search Volume", "Competition Level", "Priority",
        "Target Publish Date", "Assigned To", "Status", "Notes"
    ]
    
    for col_num, header in enumerate(gap_headers, 1):
        cell = ws7.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample gap data
    gap_data = [
        ["GAP-001", "Topic Gap", "No competitor covers [specific topic]", "keyword 1, keyword 2", "High", "Low", "P1", "Month 1, Week 1", "[Writer]", "Not Started", "Blue ocean opportunity"],
        ["GAP-002", "Format Gap", "No competitor creates video content", "video keywords", "Medium", "Low", "P2", "Month 2, Week 1", "[Video Team]", "Not Started", "Differentiation opportunity"],
        ["GAP-003", "Depth Gap", "All competitor content is superficial", "comprehensive guide keywords", "High", "Medium", "P1", "Month 1, Week 2", "[Writer]", "Not Started", "Create ultimate guide"]
    ]
    
    for row_num, row_data in enumerate(gap_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws7.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Column widths
    gap_col_widths = [12, 15, 40, 30, 18, 18, 10, 20, 20, 15, 35]
    for col_num, width in enumerate(gap_col_widths, 1):
        ws7.column_dimensions[get_column_letter(col_num)].width = width
    
    ws7.freeze_panes = 'A2'
    
    # Save the workbook
    wb.save(filename)
    print(f"✅ Enhanced Content Plan spreadsheet created: {filename}")
    return filename

if __name__ == "__main__":
    create_enhanced_content_plan()
