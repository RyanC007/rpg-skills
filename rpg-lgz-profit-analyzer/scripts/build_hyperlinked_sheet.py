"""
build_hyperlinked_sheet.py
--------------------------
Reads the master spreadsheet (built by build_master.py) and adds:
  - Hyperlinks on the ORDER # column → Google Drive PO folder
  - Hyperlinks on the Tracking # column → UPS tracking page

Requires:
  - /tmp/drive_folder_map.json (from get_drive_folders.py)
  - The master spreadsheet at /home/ubuntu/upload/Logoclothz_*_MASTER.xlsx

Usage: python3 build_hyperlinked_sheet.py [optional: path to master xlsx]

Output: /home/ubuntu/upload/Logoclothz_*_LINKED.xlsx
"""

import sys
import glob
import json
import openpyxl
from openpyxl.styles import Font

DRIVE_MAP_PATH = '/tmp/drive_folder_map.json'
RCLONE_CONFIG  = '/home/ubuntu/.gdrive-rclone.ini'

# Find the master spreadsheet
if len(sys.argv) > 1:
    master_path = sys.argv[1]
else:
    candidates = sorted(glob.glob('/home/ubuntu/upload/Logoclothz_*_MASTER.xlsx'))
    if not candidates:
        print("ERROR: No master spreadsheet found in /home/ubuntu/upload/")
        print("Run build_master.py first, or pass the path as an argument.")
        exit(1)
    master_path = candidates[-1]  # Most recent

out_path = master_path.replace('_MASTER.xlsx', '_LINKED.xlsx')
print(f"Input:  {master_path}")
print(f"Output: {out_path}")

# Load Drive folder map
if not __import__('os').path.exists(DRIVE_MAP_PATH):
    print("ERROR: /tmp/drive_folder_map.json not found. Run get_drive_folders.py first.")
    exit(1)

with open(DRIVE_MAP_PATH) as f:
    drive_map = json.load(f)  # Keys are UPPERCASE folder names

def get_drive_url(po_value):
    if not po_value or str(po_value).strip() in ('', 'nan', 'ORDER #', 'TOTALS'):
        return None
    clean = str(po_value).strip().replace('P.O. ', '').replace('P.O.', '').strip().upper()
    return drive_map.get(clean)

def get_ups_url(tracking):
    if not tracking or str(tracking).strip() in ('', 'nan', 'N/A', 'Tracking #'):
        return None
    t = str(tracking).strip()
    if len(t) >= 10:
        return f'https://www.ups.com/track?tracknum={t}'
    return None

# Load workbook
wb = openpyxl.load_workbook(master_path)
ws = wb['Order Detail']

# Find header row and column positions
header_row_idx = None
order_col = tracking_col = None

for row in ws.iter_rows():
    for cell in row:
        val = str(cell.value or '').strip()
        if val == 'ORDER #':
            header_row_idx = cell.row
            order_col = cell.column
        if val == 'Tracking #':
            tracking_col = cell.column
    if header_row_idx:
        break

if not header_row_idx:
    print("ERROR: Could not find 'ORDER #' header in the spreadsheet.")
    exit(1)

print(f"Header row: {header_row_idx}, ORDER # col: {order_col}, Tracking # col: {tracking_col}")

# Font styles
CYAN_LINK  = Font(color='00CEC9', underline='single', bold=True, size=10)
BLUE_LINK  = Font(color='74B9FF', underline='single', size=9)
WHITE_BOLD = Font(color='FFFFFF', bold=True, size=10)
GRAY_TEXT  = Font(color='B2BEC3', size=9)

matched_po = unmatched_po = linked_track = 0

for row in ws.iter_rows(min_row=header_row_idx + 1):
    # ORDER # hyperlink
    if order_col:
        po_cell = row[order_col - 1]
        url = get_drive_url(po_cell.value)
        if url:
            po_cell.hyperlink = url
            po_cell.font = CYAN_LINK
            matched_po += 1
        elif po_cell.value and str(po_cell.value).strip() not in ('', 'nan', 'TOTALS'):
            po_cell.font = WHITE_BOLD
            unmatched_po += 1

    # Tracking # hyperlink
    if tracking_col:
        track_cell = row[tracking_col - 1]
        url = get_ups_url(track_cell.value)
        if url:
            track_cell.hyperlink = url
            track_cell.font = BLUE_LINK
            linked_track += 1
        elif track_cell.value and str(track_cell.value).strip() not in ('', 'nan'):
            track_cell.font = GRAY_TEXT

wb.save(out_path)

print(f"\nResults:")
print(f"  ORDER # linked to Drive:  {matched_po}")
print(f"  ORDER # no folder found:  {unmatched_po}")
print(f"  Tracking # linked to UPS: {linked_track}")
print(f"\nSaved: {out_path}")
