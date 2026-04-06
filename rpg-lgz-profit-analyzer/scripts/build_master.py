"""
build_master.py
---------------
Builds the Logoclothz monthly real-profit master spreadsheet.

BEFORE RUNNING: Update the CONFIG section below with the correct file paths
for the month being analyzed.

Output: /home/ubuntu/upload/Logoclothz_{MONTH}_RealProfit_MASTER.xlsx
"""

import pandas as pd
import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
# CONFIG — Update these paths for each month
# ═══════════════════════════════════════════════════════════════
SALES_CSV_PATH    = '/home/ubuntu/upload/CHANGE_ME_sales.csv'
BRAINTREE_CSV     = '/home/ubuntu/upload/CHANGE_ME_braintree_fees.csv'
SHIPPING_JSON     = '/tmp/shipping_costs.json'          # from parse_inxpress.py
PAYPAL_RATE_JSON  = '/tmp/paypal_rate.json'             # from parse_paypal.py (optional)
MONTH_LABEL       = 'Month2026'                         # e.g. "February2026"

# Known custom/phone PO prefixes that use PayPal (update as needed)
PAYPAL_PO_PREFIXES = [
    'MANCAMP','PRIMACURE','GABRIELNY','JPSOLUTIONS','HALCYON','FLORINHS',
    'ARINIAI','LSHADOWS','USLUMBER','FEVERTREE','CINCH','DEFY','EMAGE',
    'ENVS','HAPPYFISH','HVAC','IAT','LITTLECITY','MTPISGAH','RBOMBERGER',
    'SAWYER','SWMEDIA','UFCGYM','USF','WAYFINDER','BONYBOOTYS','DOBRIN',
    'ASC','11VODKA'
]

# Fee rates (update if PayPal rate changes)
BRAINTREE_RATE    = 0.0289
BRAINTREE_FIXED   = 0.29
DEFAULT_PAYPAL_RATE = 0.03157   # Derived from Jan/Dec 2025-26 statements

ONE_TREE_COST     = 1.00        # Per order with One Tree flag
NEW_HOPE_COST     = 1.00        # Per order with New Hope flag
# ═══════════════════════════════════════════════════════════════

# ── Load PayPal rate ──────────────────────────────────────────
paypal_rate = DEFAULT_PAYPAL_RATE
if os.path.exists(PAYPAL_RATE_JSON):
    with open(PAYPAL_RATE_JSON) as f:
        paypal_rate = json.load(f).get('paypal_rate', DEFAULT_PAYPAL_RATE)
    print(f"PayPal rate loaded: {paypal_rate*100:.3f}%")
else:
    print(f"No PayPal rate file found, using default: {paypal_rate*100:.3f}%")

# ── Load shipping costs ───────────────────────────────────────
shipping_costs = {}
if os.path.exists(SHIPPING_JSON):
    with open(SHIPPING_JSON) as f:
        shipping_costs = json.load(f)
    print(f"Shipping costs loaded: {len(shipping_costs)} POs")
else:
    print("WARNING: No shipping_costs.json found. All shipping costs will be $0.")

# ── Load Braintree fees ───────────────────────────────────────
braintree_fees = {}
if os.path.exists(BRAINTREE_CSV):
    bt = pd.read_csv(BRAINTREE_CSV)
    # Normalize column names
    bt.columns = [c.strip() for c in bt.columns]
    # Try to find order ID and fee columns
    id_col   = next((c for c in bt.columns if 'order' in c.lower() or 'id' in c.lower()), None)
    fee_col  = next((c for c in bt.columns if 'fee' in c.lower()), None)
    cred_col = next((c for c in bt.columns if 'credit' in c.lower()), None)
    if id_col and fee_col:
        for _, row in bt.iterrows():
            po = str(row[id_col]).strip().replace('P.O. ', '').replace('P.O.', '')
            fee = abs(float(str(row[fee_col]).replace('$','').replace(',','') or 0))
            cred = abs(float(str(row.get(cred_col, 0) or 0).replace('$','').replace(',',''))) if cred_col else 0
            braintree_fees[po] = {'fee': fee, 'credit': cred}
        print(f"Braintree fees loaded: {len(braintree_fees)} transactions")
    else:
        print(f"WARNING: Could not identify columns in Braintree CSV. Columns: {list(bt.columns)}")
else:
    print("WARNING: No Braintree CSV found. Fees will be estimated.")

# ── Load sales CSV ────────────────────────────────────────────
if not os.path.exists(SALES_CSV_PATH):
    print(f"ERROR: Sales CSV not found at {SALES_CSV_PATH}")
    print("Update SALES_CSV_PATH in the CONFIG section.")
    exit(1)

raw = pd.read_csv(SALES_CSV_PATH)
raw.columns = [str(c).strip() for c in raw.columns]
print(f"Sales CSV loaded: {len(raw)} rows, columns: {list(raw.columns)}")

# ── Helper functions ──────────────────────────────────────────
def is_paypal(po):
    po_clean = str(po).replace('P.O. ','').replace('P.O.','').strip().upper()
    return any(po_clean.startswith(p) for p in PAYPAL_PO_PREFIXES)

def get_shipping(po):
    po_clean = str(po).replace('P.O. ','').replace('P.O.','').strip().upper()
    return shipping_costs.get(po_clean, 0.0)

def get_fee(po, total_revenue):
    po_clean = str(po).replace('P.O. ','').replace('P.O.','').strip().upper()
    if po_clean in braintree_fees:
        return braintree_fees[po_clean]['fee'], braintree_fees[po_clean]['credit'], 'Braintree exact'
    elif is_paypal(po):
        fee = total_revenue * paypal_rate
        return fee, 0.0, f'PayPal ({paypal_rate*100:.3f}% actual rate)'
    else:
        fee = total_revenue * BRAINTREE_RATE + BRAINTREE_FIXED
        return fee, 0.0, f'Braintree est. ({BRAINTREE_RATE*100:.2f}%+${BRAINTREE_FIXED})'

# ── Build rows ────────────────────────────────────────────────
# NOTE: This section maps column names from the sales CSV.
# Adjust the column name mapping below if the CSV uses different headers.
COLUMN_MAP = {
    'order_id':        next((c for c in raw.columns if 'order' in c.lower() and ('id' in c.lower() or '#' in c.lower() or 'num' in c.lower())), raw.columns[0]),
    'product_revenue': next((c for c in raw.columns if 'product' in c.lower() and ('total' in c.lower() or 'revenue' in c.lower() or 'paid' in c.lower())), None),
    'shipping_paid':   next((c for c in raw.columns if 'ship' in c.lower() and 'paid' in c.lower()), None),
    'taxes':           next((c for c in raw.columns if 'tax' in c.lower()), None),
    'cogs':            next((c for c in raw.columns if 'cost' in c.lower() and ('our' in c.lower() or 'cogs' in c.lower())), None),
    'tracking':        next((c for c in raw.columns if 'track' in c.lower()), None),
    'one_tree':        next((c for c in raw.columns if 'tree' in c.lower()), None),
    'new_hope':        next((c for c in raw.columns if 'hope' in c.lower()), None),
    'refund':          next((c for c in raw.columns if 'refund' in c.lower()), None),
}
print(f"\nColumn mapping: {COLUMN_MAP}")

def safe_float(val):
    try:
        return float(str(val).replace('$','').replace(',','').strip() or 0)
    except (ValueError, TypeError):
        return 0.0

def safe_str(val):
    v = str(val).strip()
    return '' if v in ('nan', 'None', 'NaN') else v

rows = []
for _, r in raw.iterrows():
    po              = safe_str(r[COLUMN_MAP['order_id']])
    product_rev     = safe_float(r[COLUMN_MAP['product_revenue']]) if COLUMN_MAP['product_revenue'] else 0
    ship_paid       = safe_float(r[COLUMN_MAP['shipping_paid']]) if COLUMN_MAP['shipping_paid'] else 0
    taxes           = safe_float(r[COLUMN_MAP['taxes']]) if COLUMN_MAP['taxes'] else 0
    cogs            = safe_float(r[COLUMN_MAP['cogs']]) if COLUMN_MAP['cogs'] else 0
    tracking        = safe_str(r[COLUMN_MAP['tracking']]) if COLUMN_MAP['tracking'] else ''
    one_tree_flag   = safe_str(r[COLUMN_MAP['one_tree']]) if COLUMN_MAP['one_tree'] else ''
    new_hope_flag   = safe_str(r[COLUMN_MAP['new_hope']]) if COLUMN_MAP['new_hope'] else ''
    refund          = safe_float(r[COLUMN_MAP['refund']]) if COLUMN_MAP['refund'] else 0

    total_rev       = product_rev + ship_paid
    ship_cost       = get_shipping(po)
    ship_margin     = ship_paid - ship_cost
    fee, credit, fee_src = get_fee(po, total_rev)
    one_tree_cost   = ONE_TREE_COST if one_tree_flag and one_tree_flag not in ('0','False','','nan') else 0
    new_hope_cost   = NEW_HOPE_COST if new_hope_flag and new_hope_flag not in ('0','False','','nan') else 0
    old_profit      = product_rev - cogs
    real_profit     = total_rev - cogs - ship_cost - fee + credit - refund - one_tree_cost - new_hope_cost
    variance        = real_profit - old_profit
    processor       = 'PayPal' if is_paypal(po) else 'Braintree'

    rows.append({
        'ORDER #':           po,
        'Processor':         processor,
        'Product Revenue':   product_rev,
        'Shipping Collected':ship_paid,
        'Taxes Collected':   taxes,
        'Total Revenue':     total_rev,
        'Our COGS':          cogs,
        'Product Margin':    product_rev - cogs,
        'InXpress Ship Cost':ship_cost,
        'Ship Source':       'InXpress ✓' if ship_cost > 0 else 'Pending',
        'Shipping Margin':   ship_margin,
        'Transaction Fee':   fee,
        'Fee Credit':        credit,
        'Fee Source':        fee_src,
        'Refund Issued':     refund,
        'One Tree Cost':     one_tree_cost,
        'New Hope Cost':     new_hope_cost,
        'REAL PROFIT':       real_profit,
        'Old Profit':        old_profit,
        'Variance':          variance,
        'Tracking #':        tracking,
    })

df = pd.DataFrame(rows)

# ── Write Excel ───────────────────────────────────────────────
out_path = f'/home/ubuntu/upload/Logoclothz_{MONTH_LABEL}_RealProfit_MASTER.xlsx'

# Use openpyxl for styled output
wb = Workbook()
ws = wb.active
ws.title = 'Order Detail'

# Color palette
DARK_BG  = '0D0D1A'; MID_BG   = '12122A'; ACC_BG   = '1A1A3E'
CYAN     = '00CEC9'; GREEN    = '00B894'; AMBER    = 'FDCB6E'
RED      = 'FF6B6B'; WHITE    = 'FFFFFF'; LGRAY    = 'B2BEC3'

def hdr_fill(hex_color): return PatternFill('solid', fgColor=hex_color)
def hdr_font(hex_color, bold=True, size=10): return Font(color=hex_color, bold=bold, size=size)
thin = Side(style='thin', color='2D3436')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Notes row
notes = (
    f"Braintree: {BRAINTREE_RATE*100:.2f}% + ${BRAINTREE_FIXED}/txn (exact from transaction report)   |   "
    f"PayPal: {paypal_rate*100:.3f}% (actual rate from statements)   |   "
    f"InXpress: weekly invoices   |   New Hope: ${NEW_HOPE_COST:.2f}/order   |   Route: N/A"
)
ws.append([notes] + [''] * (len(df.columns) - 1))
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
ws.cell(1, 1).font = Font(color=AMBER, italic=True, size=9)
ws.cell(1, 1).fill = hdr_fill(DARK_BG)

# Blank separator
ws.append([''] * len(df.columns))
ws.row_dimensions[2].height = 4

# Section header
ws.append(['LOGOCLOTHZ  ·  MONTHLY PROFIT ANALYSIS  ·  ' + MONTH_LABEL.upper()] + [''] * (len(df.columns) - 1))
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(df.columns))
ws.cell(3, 1).font = Font(color=WHITE, bold=True, size=14)
ws.cell(3, 1).fill = hdr_fill(ACC_BG)
ws.cell(3, 1).alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[3].height = 28

# Column headers
header_colors = {
    'ORDER #': CYAN, 'Processor': LGRAY,
    'Product Revenue': GREEN, 'Shipping Collected': GREEN, 'Taxes Collected': LGRAY,
    'Total Revenue': GREEN,
    'Our COGS': RED, 'Product Margin': AMBER,
    'InXpress Ship Cost': RED, 'Ship Source': LGRAY, 'Shipping Margin': AMBER,
    'Transaction Fee': RED, 'Fee Credit': GREEN, 'Fee Source': LGRAY,
    'Refund Issued': RED, 'One Tree Cost': RED, 'New Hope Cost': RED,
    'REAL PROFIT': GREEN, 'Old Profit': LGRAY, 'Variance': AMBER,
    'Tracking #': CYAN,
}
ws.append(list(df.columns))
for col_idx, col_name in enumerate(df.columns, 1):
    cell = ws.cell(4, col_idx)
    color = header_colors.get(col_name, WHITE)
    cell.font = Font(color=color, bold=True, size=9)
    cell.fill = hdr_fill(MID_BG)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[4].height = 30

# Data rows
money_cols = {'Product Revenue','Shipping Collected','Taxes Collected','Total Revenue',
              'Our COGS','Product Margin','InXpress Ship Cost','Shipping Margin',
              'Transaction Fee','Fee Credit','Refund Issued','One Tree Cost',
              'New Hope Cost','REAL PROFIT','Old Profit','Variance'}

for row_idx, (_, data_row) in enumerate(df.iterrows(), 5):
    is_paypal_row = data_row['Processor'] == 'PayPal'
    is_zero       = data_row['Product Revenue'] == 0
    row_bg = 'MID_BG' if row_idx % 2 == 0 else '0F0F22'

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row_idx, col_idx)
        val  = data_row[col_name]
        cell.value = val
        cell.border = border
        cell.alignment = Alignment(horizontal='right' if col_name in money_cols else 'left',
                                   vertical='center')

        # Background
        if is_zero:
            cell.fill = hdr_fill('1A0A0A')
        elif is_paypal_row:
            cell.fill = hdr_fill('1A1500')
        elif row_idx % 2 == 0:
            cell.fill = hdr_fill(MID_BG)
        else:
            cell.fill = hdr_fill('0F0F22')

        # Font color
        if col_name == 'REAL PROFIT':
            color = GREEN if (val or 0) >= 0 else RED
            cell.font = Font(color=color, bold=True, size=10)
        elif col_name == 'Variance':
            color = GREEN if (val or 0) >= 0 else RED
            cell.font = Font(color=color, size=9)
        elif col_name in money_cols:
            cell.font = Font(color=WHITE, size=9)
            if col_name in money_cols:
                cell.number_format = '$#,##0.00'
        elif col_name == 'ORDER #':
            cell.font = Font(color=CYAN, bold=True, size=10)
        elif col_name == 'Tracking #':
            cell.font = Font(color=LGRAY, size=8)
        else:
            cell.font = Font(color=LGRAY, size=9)

# Totals row
totals_row = len(df) + 5
ws.cell(totals_row, 1).value = 'TOTALS'
ws.cell(totals_row, 1).font = Font(color=AMBER, bold=True, size=10)
ws.cell(totals_row, 1).fill = hdr_fill(ACC_BG)

for col_idx, col_name in enumerate(df.columns, 1):
    if col_name in money_cols:
        total = df[col_name].sum()
        cell = ws.cell(totals_row, col_idx)
        cell.value = total
        cell.number_format = '$#,##0.00'
        cell.font = Font(color=AMBER, bold=True, size=10)
        cell.fill = hdr_fill(ACC_BG)
        cell.border = border

# Column widths
col_widths = {
    'ORDER #': 18, 'Processor': 12, 'Product Revenue': 14, 'Shipping Collected': 14,
    'Taxes Collected': 12, 'Total Revenue': 14, 'Our COGS': 12, 'Product Margin': 13,
    'InXpress Ship Cost': 14, 'Ship Source': 14, 'Shipping Margin': 13,
    'Transaction Fee': 13, 'Fee Credit': 10, 'Fee Source': 32,
    'Refund Issued': 11, 'One Tree Cost': 11, 'New Hope Cost': 11,
    'REAL PROFIT': 14, 'Old Profit': 12, 'Variance': 11, 'Tracking #': 24,
}
for col_idx, col_name in enumerate(df.columns, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 12)

# Freeze panes
ws.freeze_panes = 'B5'

# ── Summary tab ───────────────────────────────────────────────
ws2 = wb.create_sheet('Summary')
ws2.sheet_view.showGridLines = False

active = df[df['Product Revenue'] > 0]
total_rev    = active['Total Revenue'].sum()
total_cogs   = active['Our COGS'].sum()
total_ship   = active['InXpress Ship Cost'].sum()
total_fees   = active['Transaction Fee'].sum()
total_profit = active['REAL PROFIT'].sum()
margin_pct   = total_profit / total_rev * 100 if total_rev > 0 else 0

summary_data = [
    ('LOGOCLOTHZ  ·  MONTHLY PROFIT SUMMARY  ·  ' + MONTH_LABEL.upper(), '', ''),
    ('', '', ''),
    ('Metric', 'Amount', 'Notes'),
    ('Total Orders (active)', len(active), ''),
    ('Total Revenue', total_rev, 'Product + Shipping collected'),
    ('Our COGS', -total_cogs, 'Production cost'),
    ('InXpress Shipping', -total_ship, 'Actual shipping cost'),
    ('Transaction Fees', -total_fees, 'Braintree + PayPal'),
    ('Other Costs', -(active['Refund Issued'].sum() + active['One Tree Cost'].sum() + active['New Hope Cost'].sum()), 'Refunds + Donations'),
    ('REAL NET PROFIT', total_profit, f'{margin_pct:.1f}% net margin'),
    ('', '', ''),
    ('Old Recorded Profit', active['Old Profit'].sum(), 'Before cost enrichment'),
    ('Overstatement', active['Old Profit'].sum() - total_profit, 'Difference'),
]

for row_data in summary_data:
    ws2.append(list(row_data))

for cell in ws2[1]:
    cell.font = Font(color=WHITE, bold=True, size=14)
    cell.fill = hdr_fill(ACC_BG)
ws2.merge_cells('A1:C1')
ws2.row_dimensions[1].height = 28

for cell in ws2[3]:
    cell.font = Font(color=AMBER, bold=True, size=10)
    cell.fill = hdr_fill(MID_BG)

for row_idx in range(4, len(summary_data) + 1):
    for cell in ws2[row_idx]:
        cell.fill = hdr_fill(DARK_BG if row_idx % 2 == 0 else '0F0F22')
        cell.font = Font(color=WHITE, size=10)
    # Highlight profit row
    if ws2.cell(row_idx, 1).value == 'REAL NET PROFIT':
        for cell in ws2[row_idx]:
            cell.font = Font(color=GREEN, bold=True, size=12)
            cell.fill = hdr_fill('0D2818')

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 35

wb.save(out_path)
print(f"\nSaved: {out_path}")
print("=" * 60)
print(f"  REAL NET PROFIT:   $ {total_profit:,.2f}")
print(f"  Old Recorded:      $ {active['Old Profit'].sum():,.2f}")
print(f"  Overstatement:     $ {active['Old Profit'].sum() - total_profit:,.2f}")
print(f"  Profit Margin:     {margin_pct:.1f}%")
print("=" * 60)
