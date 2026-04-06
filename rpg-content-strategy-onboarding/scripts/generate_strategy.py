#!/usr/bin/env python3
"""
RPG Content Strategy Onboarding Engine
Ingests a standardized keyword research spreadsheet and generates a
structured content strategy document and agent routing README.

Usage:
    python3 generate_strategy.py \
        --client "Client Name" \
        --spreadsheet /path/to/keyword_research.xlsx \
        --output /tmp/content_strategy_output \
        [--brand-pillars "Pillar 1,Pillar 2,Pillar 3,Pillar 4"] \
        [--audiences "Homeowner,Investor"] \
        [--start-date "2026-04-01"]
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("sudo pip3 install openpyxl python-dateutil -q")
    import openpyxl

try:
    from dateutil.relativedelta import relativedelta
except ImportError:
    os.system("sudo pip3 install python-dateutil -q")
    from dateutil.relativedelta import relativedelta


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS & DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_BRAND_PILLARS = [
    "Decision Guidance",
    "Plan Protection",
    "Craftsmanship That Holds Up",
    "A Warm, Family-Level Experience",
]

PILLAR_KEYWORD_MAP = {
    "Decision Guidance": ["cost", "how much", "choose", "best", "what to look", "save money", "finance", "value", "tips", "guide"],
    "Plan Protection": ["permit", "prepare", "timeline", "how long", "process", "checklist", "plan", "schedule", "avoid"],
    "Craftsmanship That Holds Up": ["quality", "material", "durable", "last", "warranty", "craftsmanship", "build"],
    "A Warm, Family-Level Experience": ["experience", "trust", "family", "local", "community", "relationship", "story"],
}

INVESTOR_KEYWORDS = [
    "investor", "flip", "flipping", "rental", "rehab", "turnkey", "roi",
    "investment", "property", "wholesale", "fix and flip", "house flip",
]


def infer_audience(keyword: str) -> str:
    kw_lower = keyword.lower()
    for inv_kw in INVESTOR_KEYWORDS:
        if inv_kw in kw_lower:
            return "Investor"
    return "Homeowner"


def infer_brand_pillar(keyword: str, pillars: list) -> str:
    kw_lower = keyword.lower()
    for pillar, trigger_words in PILLAR_KEYWORD_MAP.items():
        if pillar in pillars:
            for word in trigger_words:
                if word in kw_lower:
                    return pillar
    return pillars[0] if pillars else "Decision Guidance"


def keyword_to_title(keyword: str) -> str:
    """Convert a question keyword to a blog-ready title."""
    kw = keyword.strip()
    # Capitalize first letter of each major word
    kw = kw[0].upper() + kw[1:]
    # Add location context if missing and it's a local keyword
    return kw


def parse_volume(vol_str) -> int:
    """Parse volume strings like '2200', '300-700', '50-100' into a sortable integer."""
    if vol_str is None:
        return 0
    s = str(vol_str).strip().replace(",", "")
    if "-" in s:
        parts = s.split("-")
        try:
            return int(parts[0])
        except ValueError:
            return 0
    try:
        return int(s)
    except ValueError:
        return 0


def load_aeo_questions(wb) -> list:
    """Load and rank blog topics from the AEO Questions sheet."""
    topics = []
    if "AEO Questions" not in wb.sheetnames:
        return topics

    ws = wb["AEO Questions"]
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val:
            headers[str(val).strip()] = col

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[0]:
            continue
        keyword = str(row[0]).strip()
        volume = parse_volume(row[1] if len(row) > 1 else 0)
        opportunity = str(row[6]).strip().upper() if len(row) > 6 and row[6] else "MEDIUM"
        content_type = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        topics.append({
            "keyword": keyword,
            "volume": volume,
            "opportunity": opportunity,
            "content_type": content_type,
            "source": "AEO",
        })

    # Sort: HIGH opportunity first, then by volume descending
    priority_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    topics.sort(key=lambda x: (priority_order.get(x["opportunity"], 1), -x["volume"]))
    return topics


def load_blue_ocean_topics(wb) -> list:
    """Load Blue Ocean investor topics from the Investor Flipping sheet."""
    topics = []
    sheet_name = None
    for name in wb.sheetnames:
        if "investor" in name.lower() or "blue ocean" in name.lower() or "flipping" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        return topics

    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row[0]:
            continue
        keyword = str(row[0]).strip()
        opportunity_type = str(row[6]).strip().upper() if len(row) > 6 and row[6] else ""
        if "BLUE OCEAN" in opportunity_type or "NICHE" in opportunity_type:
            topics.append({
                "keyword": keyword,
                "volume": parse_volume(row[1] if len(row) > 1 else 0),
                "opportunity": "HIGH",
                "content_type": "Investor Guide",
                "source": "BLUE_OCEAN",
            })
    return topics[:3]  # Cap at 3 Blue Ocean posts


def generate_publish_dates(count: int, start_date_str: str) -> list:
    """Generate alternating 1st and 15th publish dates starting from start_date."""
    try:
        start = datetime.strptime(start_date_str, "%Y-%m-%d")
    except ValueError:
        # Default to first of next month
        today = datetime.today()
        start = (today.replace(day=1) + relativedelta(months=1))

    dates = []
    current = start
    for i in range(count):
        dates.append(current.strftime("%Y-%m-%d"))
        if current.day == 1:
            current = current.replace(day=15)
        else:
            current = (current.replace(day=1) + relativedelta(months=1))
    return dates


def build_calendar(aeo_topics: list, blue_ocean_topics: list, pillars: list, start_date: str) -> list:
    """Combine topics into a 12-post calendar."""
    # Interleave: 2 homeowner AEO posts, 1 investor Blue Ocean post, repeat
    calendar_topics = []
    aeo_idx = 0
    bo_idx = 0

    while len(calendar_topics) < 12:
        # Add 2 AEO (homeowner-leaning) posts
        for _ in range(2):
            if aeo_idx < len(aeo_topics) and len(calendar_topics) < 12:
                calendar_topics.append(aeo_topics[aeo_idx])
                aeo_idx += 1
        # Add 1 Blue Ocean (investor) post if available, else another AEO
        if bo_idx < len(blue_ocean_topics) and len(calendar_topics) < 12:
            calendar_topics.append(blue_ocean_topics[bo_idx])
            bo_idx += 1
        elif aeo_idx < len(aeo_topics) and len(calendar_topics) < 12:
            calendar_topics.append(aeo_topics[aeo_idx])
            aeo_idx += 1
        else:
            break

    dates = generate_publish_dates(len(calendar_topics), start_date)

    calendar = []
    for i, (topic, date) in enumerate(zip(calendar_topics, dates), 1):
        audience = infer_audience(topic["keyword"])
        pillar = infer_brand_pillar(topic["keyword"], pillars)
        title = keyword_to_title(topic["keyword"])
        calendar.append({
            "num": i,
            "publish_date": date,
            "title": title,
            "keyword": topic["keyword"],
            "audience": audience,
            "pillar": pillar,
            "source": topic["source"],
            "status": "To Be Written",
        })

    return calendar


def render_calendar_table(calendar: list) -> str:
    """Render the calendar as a Markdown table."""
    header = "| # | Publish Date | Blog Title | Primary Keyword | Target Audience | Brand Pillar | Status |\n"
    separator = "|---|---|---|---|---|---|---|\n"
    rows = ""
    for post in calendar:
        rows += f"| {post['num']} | {post['publish_date']} | {post['title']} | `{post['keyword']}` | {post['audience']} | {post['pillar']} | {post['status']} |\n"
    return header + separator + rows


def generate_strategy_doc(client_name: str, calendar: list, pillars: list, audiences: list, output_dir: str):
    """Write the full Content Strategy Markdown document."""
    client_slug = re.sub(r"[^a-z0-9]+", "_", client_name.lower()).strip("_")
    filename = f"{client_slug}_Content_Strategy_v1.md"
    filepath = os.path.join(output_dir, filename)

    calendar_table = render_calendar_table(calendar)
    pillar_list = "\n".join([f"{i+1}. **{p}**" for i, p in enumerate(pillars)])
    today = datetime.today().strftime("%B %d, %Y")

    content = f"""# {client_name} — Content Strategy & Blog Plan v1
**Status:** Active | **Version:** 1.0 | **Generated:** {today}

---

## 1. Core Content Mission

All content produced for {client_name} must build trust, demonstrate expertise, and drive qualified leads. Every piece of content must align with one of the four brand pillars:

{pillar_list}

---

## 2. Target Audiences & Keyword Strategy

Content is segmented to address the unique pain points and motivators of each primary audience: **{" and ".join(audiences)}**.

Audience definitions, pain points, and keyword assignments are maintained in the client's Master Keyword Research spreadsheet. The primary keywords for each audience are drawn from the `Keyword Pool` tab of that spreadsheet.

---

## 3. Blog Content Calendar (First 12 Posts)

This calendar is derived from the **Master Keyword Research** spreadsheet (AEO Questions & Blue Ocean tabs). The Blog Content Manager agent is tasked with writing and staging these posts in order of publish date.

{calendar_table}

---

## 4. Source of Truth & Agent Instructions

This document is the **definitive source of truth** for all {client_name} content strategy. It is located in the project shared files and loads automatically for all tasks within this project.

| Agent | Instruction |
| :--- | :--- |
| **Blog Content Manager** | Pull blog titles, keywords, and audience from the calendar above. Write posts in order of publish date. Stage completed posts in the `Content_Pipeline` folder on Google Drive. |
| **WordPress Publishing Agent** | Confirm a post appears in this calendar with `Status: Approved` before publishing to the site. |
| **LinkedIn Manager** | Align all LinkedIn posts with the brand pillars and audience pain points defined in this document. |
| **SEO / Analytics Manager** | Track keyword rankings for all primary keywords listed in the calendar above. Report on ranking progress monthly. |

---

## 5. Version Control

When a new content plan is received (e.g., from the client or their marketing team), merge it with this document and save the updated version as `{client_slug}_Content_Strategy_v2.md` in the same folder. Update the agent instructions in the `Content_Plan/README.md` to point to the new version.
"""

    with open(filepath, "w") as f:
        f.write(content)

    print(f"✅ Strategy document written: {filepath}")
    return filename


def generate_readme(client_name: str, strategy_filename: str, gdrive_path: str, github_path: str, output_dir: str):
    """Write the Content_Plan folder README for agent routing."""
    filepath = os.path.join(output_dir, "README.md")
    client_slug = re.sub(r"[^a-z0-9]+", "_", client_name.lower()).strip("_")
    today = datetime.today().strftime("%B %d, %Y")

    content = f"""# Content Plan — {client_name}

**Status:** ACTIVE | **Last Updated:** {today}

## Primary Document

**`{strategy_filename}`** is the definitive source of truth for all {client_name} content strategy and blog calendar.

| Location | Path |
| :--- | :--- |
| **Google Drive** | `{gdrive_path}/{strategy_filename}` |
| **GitHub** | `{github_path}/{strategy_filename}` |
| **Project Shared Files** | Loaded automatically in every task for this project |

## Agent Instructions

All agents in the AI Manager Stack MUST read `{strategy_filename}` at the start of any content-related task.

| Agent | Instruction |
| :--- | :--- |
| **Blog Content Manager** | Pull blog titles, keywords, and audience from the content calendar. Write posts in order of publish date. |
| **WordPress Publishing Agent** | Confirm post is approved in the calendar before publishing. |
| **LinkedIn Manager** | Align all posts with brand pillars and audiences defined in the strategy document. |
| **SEO / Analytics Manager** | Track all primary keywords listed in the strategy document. |

## Updating This Plan

When a new content plan is received, save the updated strategy as `{client_slug}_Content_Strategy_v2.md` in this same folder and update the paths in this README.
"""

    with open(filepath, "w") as f:
        f.write(content)

    print(f"✅ README written: {filepath}")


def main():
    parser = argparse.ArgumentParser(description="RPG Content Strategy Onboarding Engine")
    parser.add_argument("--client", required=True, help="Client name (e.g., 'Red Horse Construction')")
    parser.add_argument("--spreadsheet", required=True, help="Path to keyword research spreadsheet (.xlsx)")
    parser.add_argument("--output", default="/tmp/content_strategy_output", help="Output directory")
    parser.add_argument("--brand-pillars", default=None, help="Comma-separated brand pillars")
    parser.add_argument("--audiences", default="Homeowner,Investor", help="Comma-separated target audiences")
    parser.add_argument("--start-date", default=None, help="First publish date (YYYY-MM-DD). Defaults to next month.")
    parser.add_argument("--gdrive-path", default="Content_Plan", help="Google Drive path for agent routing README")
    parser.add_argument("--github-path", default="clients/content", help="GitHub path for agent routing README")
    args = parser.parse_args()

    # Setup
    os.makedirs(args.output, exist_ok=True)
    pillars = args.brand_pillars.split(",") if args.brand_pillars else DEFAULT_BRAND_PILLARS
    audiences = args.audiences.split(",")

    # Determine start date
    if args.start_date:
        start_date = args.start_date
    else:
        today = datetime.today()
        start_date = (today.replace(day=1) + relativedelta(months=1)).strftime("%Y-%m-%d")

    print(f"\n🚀 RPG Content Strategy Onboarding Engine")
    print(f"   Client: {args.client}")
    print(f"   Spreadsheet: {args.spreadsheet}")
    print(f"   Output: {args.output}")
    print(f"   Start Date: {start_date}\n")

    # Load spreadsheet
    print("📊 Loading keyword spreadsheet...")
    wb = openpyxl.load_workbook(args.spreadsheet)
    print(f"   Sheets found: {wb.sheetnames}")

    # Extract topics
    print("🔍 Extracting AEO question topics...")
    aeo_topics = load_aeo_questions(wb)
    print(f"   Found {len(aeo_topics)} AEO topics")

    print("🔵 Extracting Blue Ocean investor topics...")
    blue_ocean_topics = load_blue_ocean_topics(wb)
    print(f"   Found {len(blue_ocean_topics)} Blue Ocean topics")

    # Build calendar
    print("📅 Building 12-post content calendar...")
    calendar = build_calendar(aeo_topics, blue_ocean_topics, pillars, start_date)
    print(f"   Calendar built with {len(calendar)} posts")

    # Generate files
    print("📝 Generating strategy document...")
    strategy_filename = generate_strategy_doc(args.client, calendar, pillars, audiences, args.output)

    print("📋 Generating agent routing README...")
    generate_readme(args.client, strategy_filename, args.gdrive_path, args.github_path, args.output)

    # Output summary JSON for the calling agent
    summary = {
        "client": args.client,
        "strategy_file": strategy_filename,
        "output_dir": args.output,
        "post_count": len(calendar),
        "start_date": start_date,
        "files_generated": [strategy_filename, "README.md"],
    }
    summary_path = os.path.join(args.output, "onboarding_summary.json")
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)

    print(f"\n✅ Onboarding complete. Summary: {summary_path}")
    print(f"   Files ready in: {args.output}")
    print(f"   Next step: Deploy files to Google Drive and GitHub using the SKILL.md deployment instructions.")


if __name__ == "__main__":
    main()
